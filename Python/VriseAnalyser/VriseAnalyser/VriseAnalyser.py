from vriseScrapper import Vrise
import pandas as pd
from openpyxl import load_workbook, Workbook
import os
import pickle

cellNumberPath = 'cellNumber'
vrise_url = r'VRise_options.html'
excel_url = 'vrise.xlsx'

def saveCellNumber(cellNumber):
    try:
        p_out = open(cellNumberPath, 'wb')
        pickle.dump(cellNumber, p_out)
        p_out.close()
    except:
        print( 'Could not save cell number' )

def readCellNumber():
    try:
        p_in = open(cellNumberPath, 'rb')
        result = pickle.load(p_in)
        return result
    except EOFError as error:
        print( 'Could not write cell number' )

v = Vrise(vrise_url)
dVar = v.getDailyVariations()

if not os.path.exists(cellNumberPath):
    saveCellNumber(2)

if not os.path.exists(excel_url):
    wb = Workbook()
    ws = wb.active
    i = 0
    ch = 'B'
    headings = v.getLabels()

    for h in headings:
        col = ord(ch) + i

        if not col > ord('Z'):
            ws[chr(col) + str(1)] = h
            i += 1
        else:
            break

    wb.save(excel_url)
wb = load_workbook(excel_url)
cellNumber = readCellNumber()
ws = wb.active
i = 0
ch = 'B'

for v in dVar:
    col = ord(ch) + i

    if not col > ord('Z'):
        ws[chr(col) + str(cellNumber)] = v
        i += 1
    else:
        break

cellNumber += 1
saveCellNumber(cellNumber)
wb.save(excel_url)