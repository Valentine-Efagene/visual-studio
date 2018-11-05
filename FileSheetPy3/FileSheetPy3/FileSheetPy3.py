from openpyxl import load_workbook
import pickle
import os
import time

cellNumberPath = 'cellNumber'

def saveCellNumber(cellNumber):
    try:
        p_out = open(cellNumberPath, 'wb')
        pickle.dump(cellNumber, p_out)
        p_out.close()
    except:
        print ('Could not save cell number')

def readCellNumber():
    try:
        p_in = open(cellNumberPath, 'rb')
        result = pickle.load(p_in)
        return result
    except EOFError as error:
        print ('Could not write cell number')

path = 'C:/Users/valentyne/Documents/test/'

if not os.path.exists(cellNumberPath):
    saveCellNumber(2)

list_of_files = os.listdir(path)
length = len(list_of_files)
wb = load_workbook("C:/Users/valentyne/Documents/spreedsheets/sample.xlsx")
nameLoc = 'A'
timeLoc = 'B'
pathLoc = 'C'
cellNumber = readCellNumber()

while(True):
    list_of_files = os.listdir(path)
    paths = [os.path.join(path, basename) for basename in list_of_files]

    if (len(list_of_files) > 0) and len(list_of_files) > length:
        latest_file = max(paths, key=os.path.getctime)
        print (os.path.basename(latest_file))
        length = len(list_of_files)
        # grab the active worksheet
        ws = wb.active

        # Data can be assigned directly to cells
        ws[nameLoc+str(cellNumber)] = os.path.basename(latest_file)
        ws[timeLoc+str(cellNumber)] = time.asctime( time.localtime(os.path.getctime(latest_file)))
        ws[pathLoc+str(cellNumber)].hyperlink = latest_file
        ws.value = latest_file
        ws.style = "Hyperlink"
        cellNumber += 1
        saveCellNumber(cellNumber)
        # Save the file
        wb.save("C:/Users/valentyne/Documents/spreedsheets/sample.xlsx")
